import pdb
import time
import json
import os
import random
import re
from multiprocessing import Pool
import multiprocessing
from transformers import AutoTokenizer, AutoModelForCausalLM
import rich.progress

from tqdm import tqdm
import utils

import openpyxl

random.seed(42)

def encode_prompt(prompt_instructions, prompt):
    prompt = prompt + "\n"
    
    idx = 1
    for _, task_dict in enumerate(prompt_instructions):
        (instruction, input, output) = task_dict["instruction"], task_dict["input"], task_dict["output"]
        instruction = re.sub(r"\s+", " ", instruction).strip().rstrip(":")
        input = "<noinput>" if input == "" else input
        prompt += f"###\n"
        prompt += f"[{idx}]. 指令:{instruction}\n"
        prompt += f"[{idx}]. 输入:\n{input}\n"
        prompt += f"[{idx}]. 输出:\n{output}\n"
        idx += 1
    prompt += f"###\n"
    prompt += f"[{idx}]. 指令:"
    return prompt

def merge_data():
    # 列出所有的文件
    files = os.listdir('./../data/zh_law_instruction/')
    print(files)
    datas = []
    for file in tqdm(files):
        if file.endswith('.json'):
            with open('./../data/zh_law_instruction/' + file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                for line in tqdm(lines):
                    data = json.loads(line)
                    datas.append(data)
    with open('./../data/zh_law_instruction/raw_zh_law_instruction.json', 'w', encoding='utf-8') as f:
        for data in tqdm(datas):
            f.write(json.dumps(data, ensure_ascii=False) + '\n')

def post_process_gpt3_response(num_prompt_instructions, response):
    if response is None:
        return []
    raw_instructions = f"[{num_prompt_instructions+1}]. 指令:" + response.message.content
    raw_instructions = re.split("###", raw_instructions)

    instructions = []
    for idx, inst in enumerate(raw_instructions):
        # if the decoding stops due to length, the last example is likely truncated so we discard it
        if idx == len(raw_instructions) - 1 and response["finish_reason"] == "length":
            continue
        idx = idx + num_prompt_instructions + 1
        splitted_data = re.split(f"\[{idx}\]\.\s+(指令|输入|输出):", inst)

        # print(splitted_data)
        if len(splitted_data) != 7:
            continue
        else:
            inst = splitted_data[2].strip()
            input = splitted_data[4].strip()
            input = "" if input.lower() == "<noinput>" else input
            output = splitted_data[6].strip()
        instructions.append({"instruction": inst, "input": input, "output": output})
    return instructions


def read_xlsx(path):
    book = openpyxl.load_workbook(path)
    sheet = book.active
    data = []
    for row in range(2, 204):
        instruction = sheet.cell(row=row, column=2).value
        input = sheet.cell(row=row, column=3).value
        output = sheet.cell(row=row, column=4).value
        if instruction != "":
            data.append({"instruction": instruction, "input": input, "output": output})
    return data

def genInstruction(cpu_id, num_instructions_to_generate, seed_instruction_data, num_prompt_instructions, output_dir):

    prompt = open("./prompt.txt", encoding='utf-8').read()
    random.seed(1024+cpu_id)
    for i in tqdm(range(num_instructions_to_generate)):
        # 随机选取num_prompt_instructions个seed instruction
        prompt_instructions = random.sample(seed_instruction_data, num_prompt_instructions)
        inputs = encode_prompt(prompt_instructions, prompt)
        batch_inputs = [inputs]
        decoding_args = utils.OpenAIDecodingArguments()
        request_start = time.time()
        results = utils.openai_completion( # 一个输入
            prompts=batch_inputs,
            batch_size=1,
            decoding_args=decoding_args,
            # logit_bias={"50256": -100},  # prevent the <|endoftext|> token from being generated
        )

        request_duration = time.time() - request_start

        process_start = time.time()
        instruction_data = []
        for result in results:
            new_instructions = post_process_gpt3_response(num_prompt_instructions, result)
            instruction_data += new_instructions
            # print(instruction_data)

        process_duration = time.time() - process_start
        print(f"Request {i} took {request_duration:.2f}s, processing took {process_duration:.2f}s")
        # print(f"Generated {total} instructions, kept {keep} instructions")
        # 写到json文件，一行一个数据
        with open(os.path.join(output_dir, f"{cpu_id+20}.json"), mode='a', encoding='utf-8') as f:
            if len(instruction_data) > 0:
                f.write('\n'.join(json.dumps(instruction, ensure_ascii=False) for instruction in instruction_data) +'\n')

        # utils.jdump(instruction_data, os.path.join(output_dir, "regen.json"))

def generate_instruction_following_data(
    output_dir="../data/",
    seed_tasks_path="../data/seed_task.xlsx",
    num_instructions_to_generate=4000,
    num_prompt_instructions=3,
    request_batch_size=1,
    num_cpus=12,
):

    seed_instruction_data = read_xlsx(seed_tasks_path)

    print(f"Loaded {len(seed_instruction_data)} seed tasks")

    for i in range(num_cpus):
        p = multiprocessing.Process(target=genInstruction, args=(i, num_instructions_to_generate, seed_instruction_data, num_prompt_instructions, output_dir))
        p.start()

if __name__ == "__main__":
    generate_instruction_following_data()

    merge_data()